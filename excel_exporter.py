"""
excel_exporter.py
Génération et mise à jour du fichier Excel de synthèse.
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import pandas as pd

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# COLONNES DU FICHIER EXCEL
# ─────────────────────────────────────────────

COLUMNS = [
    "Fichier",
    "N° Police",
    "Assuré",
    "Assureur",
    "Courtier",
    "Taux de prime",
    "Prime provisionnelle",
    "Pourcentage assuré",
    "Délai d'indemnisation",
    "Délai max crédit",
    "Date de prise d'effet",
    "Date d'échéance",
    "Durée",
    "Devise",
    "Limite de décaissement",
    "Zone discrétionnaire",
    "Groupe (nb polices)",
    "Score confiance moyen",
    "Date d'analyse",
]

FIELD_MAP = {
    "numero_police":        "N° Police",
    "assure":               "Assuré",
    "assureur":             "Assureur",
    "courtier":             "Courtier",
    "taux_prime":           "Taux de prime",
    "prime_provisionnelle": "Prime provisionnelle",
    "quotites_garanties":   "Pourcentage assuré",
    "delai_indemnisation":  "Délai d'indemnisation",
    "delai_max_credit":     "Délai max crédit",
    "date_prise_effet":     "Date de prise d'effet",
    "date_echeance":        "Date d'échéance",
    "duree_police":         "Durée",
    "devise":               "Devise",
    "limite_decaissement":  "Limite de décaissement",
    "zone_discretionnaire": "Zone discrétionnaire",
}


# ─────────────────────────────────────────────
# CONVERSION RÉSULTATS → LIGNE DATAFRAME
# ─────────────────────────────────────────────

def results_to_row(filename: str, merged_results: dict) -> dict:
    """
    Convertit les résultats d'analyse en ligne pour le DataFrame.
    merged_results : dict {field: {value, confidence, method}}
    """
    row = {"Fichier": filename}

    confidences = []
    for field, col_name in FIELD_MAP.items():
        if field == "fichier":
            continue
        field_data = merged_results.get(field, {"value": "Non trouvé", "confidence": 0.0})
        val = field_data.get("value", "Non trouvé")
        row[col_name] = val if not isinstance(val, list) else str(val)
        confidences.append(field_data.get("confidence", 0.0))

    # Groupe de polices — nombre
    gp = merged_results.get("groupe_polices", {}).get("value")
    row["Groupe (nb polices)"] = len(gp) if isinstance(gp, list) else 0

    # Score moyen (uniquement sur les champs trouvés)
    found = [c for c in confidences if c > 0]
    row["Score confiance moyen"] = f"{(sum(found)/len(found)*100):.0f}%" if found else "0%"
    row["Date d'analyse"] = datetime.now().strftime("%d/%m/%Y %H:%M")

    return row


# ─────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────

def export_to_excel(rows: List[dict], output_path: str) -> str:
    """
    Exporte les données vers un fichier Excel formaté.
    Si le fichier existe, écrase les données.
    """
    try:
        import openpyxl
        from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                      Side)
        from openpyxl.utils import get_column_letter

        df = pd.DataFrame(rows, columns=COLUMNS)

        # Créer le writer
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Contrats analysés")

            ws = writer.sheets["Contrats analysés"]

            # ── Formatage de l'en-tête ──
            header_fill = PatternFill("solid", fgColor="1F3864")  # Bleu foncé
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(border_style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx, col_name in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border

            # ── Formatage des données ──
            alt_fill = PatternFill("solid", fgColor="EEF2FF")  # Bleu très clair
            data_align = Alignment(vertical="center", wrap_text=True)

            for row_idx in range(2, len(rows) + 2):
                for col_idx in range(1, len(COLUMNS) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = data_align
                    cell.border = border
                    # Alternance de couleurs
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

                    # Colorisation du score de confiance
                    if col_idx == COLUMNS.index("Score confiance moyen") + 1:
                        val = cell.value or "0%"
                        pct = int(val.replace("%", "").strip()) if "%" in str(val) else 0
                        if pct >= 70:
                            cell.font = Font(color="1E8449", bold=True)  # Vert
                        elif pct >= 40:
                            cell.font = Font(color="D4AC0D", bold=True)  # Orange
                        else:
                            cell.font = Font(color="C0392B", bold=True)  # Rouge

                    # Colorisation "Non trouvé"
                    if str(cell.value) == "Non trouvé":
                        cell.font = Font(color="999999", italic=True)

            # ── Largeurs des colonnes ──
            col_widths = {
                "Fichier": 30,
                "N° Police": 15,
                "Assuré": 28,
                "Assureur": 28,
                "Courtier": 25,
                "Taux de prime": 15,
                "Prime provisionnelle": 20,
                "Pourcentage assuré": 18,
                "Délai d'indemnisation": 20,
                "Délai max crédit": 16,
                "Date de prise d'effet": 20,
                "Date d'échéance": 16,
                "Durée": 12,
                "Devise": 12,
                "Limite de décaissement": 22,
                "Zone discrétionnaire": 20,
                "Groupe (nb polices)": 18,
                "Score confiance moyen": 18,
                "Date d'analyse": 18,
            }
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

            # ── Hauteur de ligne ──
            ws.row_dimensions[1].height = 35
            for row_idx in range(2, len(rows) + 2):
                ws.row_dimensions[row_idx].height = 25

            # ── Figer la première ligne ──
            ws.freeze_panes = "A2"

            # ── Filtre automatique ──
            ws.auto_filter.ref = ws.dimensions

        logger.info(f"Excel exporté : {output_path}")
        return output_path

    except ImportError:
        logger.error("openpyxl non disponible")
        raise
    except Exception as e:
        logger.error(f"Erreur export Excel: {e}")
        raise


def append_or_update_excel(filename: str, merged_results: dict, output_path: str) -> str:
    """
    Ajoute ou met à jour une ligne dans le fichier Excel.
    Si le fichier existe, charge les données existantes et met à jour.
    """
    new_row = results_to_row(filename, merged_results)

    if os.path.exists(output_path):
        try:
            df_existing = pd.read_excel(output_path, sheet_name="Contrats analysés")
            # Supprimer l'ancienne ligne si même fichier
            df_existing = df_existing[df_existing["Fichier"] != filename]
            rows = df_existing.to_dict("records")
            rows.append(new_row)
        except Exception:
            rows = [new_row]
    else:
        rows = [new_row]

    return export_to_excel(rows, output_path)
