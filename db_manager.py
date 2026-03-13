"""
db_manager.py
Gestion de la base de données Excel (persistance des contrats analysés).

Fonctionnalités :
  - Charger la base depuis un fichier Excel existant
  - Upsert : mise à jour si N° police existe, sinon ajout
  - Supprimer une ou plusieurs lignes
  - Sauvegarder avec formatage complet
  - Historique des modifications (colonne Date MAJ)
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)

# Colonne clé pour identifier un contrat de façon unique
KEY_COLUMN = "N° Police"

# Colonnes de la base (même ordre que excel_exporter.COLUMNS)
DB_COLUMNS = [
    "N° Police",
    "Assuré",
    "Assureur",
    "Courtier",
    "Taux de prime",
    "Prime provisionnelle",
    "Pourcentage assuré",
    "Délai d'indemnisation",
    "Délai max crédit",
    "Date de prise d'effet",
    "Date d'échéance",
    "Durée",
    "Devise",
    "Limite de décaissement",
    "Zone discrétionnaire",
    "Groupe (nb polices)",
    "Score confiance moyen",
    "Fichier source",
    "Date d'analyse",
    "Date de MAJ",
]

SHEET_NAME = "Contrats analysés"


# ─────────────────────────────────────────────────────────────
# Chargement
# ─────────────────────────────────────────────────────────────

def load_db(path: str) -> pd.DataFrame:
    """
    Charge la base depuis un fichier Excel.
    Crée un DataFrame vide avec les bonnes colonnes si le fichier n'existe pas.
    """
    if os.path.exists(path):
        try:
            df = pd.read_excel(path, sheet_name=SHEET_NAME, dtype=str)
            # Ajouter les colonnes manquantes (évolution du schéma)
            for col in DB_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            df = df[DB_COLUMNS]
            df = df.fillna("")
            logger.info(f"Base chargée : {len(df)} contrats depuis {path}")
            return df
        except Exception as e:
            logger.error(f"Erreur chargement base : {e}")
            return _empty_df()
    return _empty_df()


def _empty_df() -> pd.DataFrame:
    return pd.DataFrame(columns=DB_COLUMNS)


# ─────────────────────────────────────────────────────────────
# Upsert (ajout ou mise à jour)
# ─────────────────────────────────────────────────────────────

def upsert_row(df: pd.DataFrame, row: dict) -> Tuple[pd.DataFrame, str]:
    """
    Insère ou met à jour une ligne identifiée par KEY_COLUMN.

    Retourne (df_mis_a_jour, action) où action = "updated" | "inserted".
    """
    key_val = str(row.get(KEY_COLUMN, "")).strip()

    # Normaliser : extraire seulement le numéro (ex: "1453314 USD" → "1453314 USD")
    mask = df[KEY_COLUMN].str.strip() == key_val

    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    row["Date de MAJ"] = now

    if mask.any():
        # Mise à jour de la ligne existante
        idx = df.index[mask][0]
        for col, val in row.items():
            if col in df.columns:
                df.at[idx, col] = val
        logger.info(f"Contrat {key_val} mis à jour")
        return df, "updated"
    else:
        # Nouvelle ligne
        if "Date d'analyse" not in row or not row["Date d'analyse"]:
            row["Date d'analyse"] = now
        new_row = {col: row.get(col, "") for col in DB_COLUMNS}
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        logger.info(f"Contrat {key_val} ajouté")
        return df, "inserted"


def exists_in_db(df: pd.DataFrame, numero_police: str) -> bool:
    """Vérifie si un N° police est déjà dans la base."""
    if df.empty or KEY_COLUMN not in df.columns:
        return False
    return df[KEY_COLUMN].str.strip().eq(str(numero_police).strip()).any()


# ─────────────────────────────────────────────────────────────
# Suppression
# ─────────────────────────────────────────────────────────────

def delete_rows(df: pd.DataFrame, indices: List[int]) -> pd.DataFrame:
    """Supprime les lignes aux indices donnés (index du DataFrame)."""
    df = df.drop(index=indices).reset_index(drop=True)
    logger.info(f"{len(indices)} ligne(s) supprimée(s)")
    return df


# ─────────────────────────────────────────────────────────────
# Conversion résultats d'analyse → ligne DB
# ─────────────────────────────────────────────────────────────

def merged_to_db_row(filename: str, merged: dict) -> dict:
    """
    Convertit les résultats merged (format app) en ligne pour la base.
    """
    def val(field):
        data = merged.get(field, {})
        v = data.get("value", "") if isinstance(data, dict) else ""
        return "" if v in [None, "Non trouvé"] else str(v)

    def conf():
        confs = []
        for field in merged:
            if field == "groupe_polices":
                continue
            d = merged.get(field, {})
            c = d.get("confidence", 0.0) if isinstance(d, dict) else 0.0
            if c > 0:
                confs.append(c)
        return f"{sum(confs)/len(confs)*100:.0f}%" if confs else "0%"

    gp = merged.get("groupe_polices", {})
    gp_val = gp.get("value") if isinstance(gp, dict) else None
    nb_polices = len(gp_val) if isinstance(gp_val, list) else 0

    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    return {
        "N° Police":             val("numero_police"),
        "Assuré":                val("assure"),
        "Assureur":              val("assureur"),
        "Courtier":              val("courtier"),
        "Taux de prime":         val("taux_prime"),
        "Prime provisionnelle":  val("prime_provisionnelle"),
        "Pourcentage assuré":    val("quotites_garanties"),
        "Délai d'indemnisation": val("delai_indemnisation"),
        "Délai max crédit":      val("delai_max_credit"),
        "Date de prise d'effet": val("date_prise_effet"),
        "Date d'échéance":       val("date_echeance"),
        "Durée":                 val("duree_police"),
        "Devise":                val("devise"),
        "Limite de décaissement":val("limite_decaissement"),
        "Zone discrétionnaire":  val("zone_discretionnaire"),
        "Groupe (nb polices)":   str(nb_polices),
        "Score confiance moyen": conf(),
        "Fichier source":        filename,
        "Date d'analyse":        now,
        "Date de MAJ":           now,
    }


# ─────────────────────────────────────────────────────────────
# Sauvegarde
# ─────────────────────────────────────────────────────────────

def save_db(df: pd.DataFrame, path: str) -> bool:
    """
    Sauvegarde le DataFrame dans le fichier Excel avec mise en forme.
    Retourne True si succès.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=SHEET_NAME)
            ws = writer.sheets[SHEET_NAME]

            # En-tête
            header_fill = PatternFill("solid", fgColor="1F3864")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            thin = Side(border_style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col_idx in range(1, len(DB_COLUMNS) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            # Données
            alt_fill = PatternFill("solid", fgColor="EEF2FF")
            for row_idx in range(2, len(df) + 2):
                for col_idx in range(1, len(DB_COLUMNS) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(vertical="center")
                    cell.border = border
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill
                    if str(cell.value) in ["Non trouvé", ""]:
                        cell.font = Font(color="999999", italic=True)

            # Largeurs
            widths = {
                "N° Police": 16, "Assuré": 28, "Assureur": 28, "Courtier": 24,
                "Taux de prime": 14, "Prime provisionnelle": 20,
                "Pourcentage assuré": 18, "Délai d'indemnisation": 20,
                "Délai max crédit": 16, "Date de prise d'effet": 20,
                "Date d'échéance": 16, "Durée": 10, "Devise": 12,
                "Limite de décaissement": 22, "Zone discrétionnaire": 20,
                "Groupe (nb polices)": 18, "Score confiance moyen": 16,
                "Fichier source": 30, "Date d'analyse": 17, "Date de MAJ": 17,
            }
            for col_idx, col in enumerate(DB_COLUMNS, start=1):
                ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col, 15)

            ws.row_dimensions[1].height = 32
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        logger.info(f"Base sauvegardée : {path} ({len(df)} contrats)")
        return True

    except Exception as e:
        logger.error(f"Erreur sauvegarde : {e}")
        return False
